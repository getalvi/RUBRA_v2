"""
RUBRA v7 — Always-Updated Agentic Intelligence
- Live Knowledge Engine (background RSS/API fetcher - always current)
- Hermes++ Coding Engine (encoder-inspired multi-layer code generation)
- Self-Prompt Loop (continuous self-learning from CL1 repo concept)
- Full multilingual · Smart Tutor · Vision · Exam Generator
"""
import os,sys,re,json,time,uuid,math,sqlite3,hashlib,asyncio,logging,base64,mimetypes
import threading,xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional,AsyncIterator
from datetime import datetime,timezone
import aiohttp,requests as _req
from fastapi import FastAPI,UploadFile,File,Form,BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse,JSONResponse
from pydantic import BaseModel

HERE = Path(__file__).resolve().parent
os.chdir(str(HERE))

# ═══════════════════════════════════════════════════════
#  API KEYS
# ═══════════════════════════════════════════════════════
ZAI_KEY  = os.getenv("ZAI_API_KEY",    "b4a30453455d4c5fa63d63ce32b71506.k1s8vXrPLKnr3m5l")
GROQ_KEY = os.getenv("GROQ_API_KEY",   "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")
OR_KEY   = os.getenv("OPENROUTER_KEY", "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")

ZAI_CHAT = "https://api.z.ai/api/paas/v4/chat/completions"
ZAI_CODE = "https://api.z.ai/api/coding/paas/v4/chat/completions"
GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"
OR_URL   = "https://openrouter.ai/api/v1/chat/completions"

DB_PATH    = HERE / "rubra.db"
UPLOAD_DIR = HERE / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("rubra")

# ═══════════════════════════════════════════════════════
#  PYDANTIC
# ═══════════════════════════════════════════════════════
class ChatRequest(BaseModel):
    message:str; session_id:Optional[str]=None; task_type:Optional[str]=None; mode:Optional[str]=None

class ExamRequest(BaseModel):
    subject:str; class_:str; topic:Optional[str]=None; q_count:Optional[int]=10
    type_:Optional[str]="mixed"; lang:Optional[str]="bn"

# ═══════════════════════════════════════════════════════
#  LANGUAGE DETECTION
# ═══════════════════════════════════════════════════════
BN_UNICODE=re.compile(r'[\u0980-\u09FF]')
BN_ROMAN=['ami','tumi','apni','ki','keno','kothay','ache','hobe','koro','bolo','jano','shekho',
    'likhte','bolte','dite','nao','dao','jao','aso','thako','dekho','eta','ota','eita','oita',
    'emon','onek','ektu','bhalo','kharap','shundor','kothin','shohoj','porashona','bishoy',
    'bujhi','bujhte','bujhao','korao','shekao','solve','explain','class','math','science',
    'chapter','porar','pora','korte','korbo','hoyeche','lagbe','chai','thakbe','parbo']

def detect_lang(text):
    if BN_UNICODE.search(text): return "bn"
    lower=text.lower()
    if sum(1 for w in BN_ROMAN if re.search(r'\b'+w+r'\b',lower))>=2: return "bn_roman"
    return "en"

def lang_instr(lang):
    if lang=="bn":      return "⚡ CRITICAL: Reply ENTIRELY in Bengali (বাংলা). Every word must be Bengali."
    if lang=="bn_roman":return "⚡ CRITICAL: Reply in Romanized Bengali (Banglish) — like a Bangladeshi friend texting. Mix Bangla vocab in English letters. Example: 'Haan, eta ekta important concept. Chalo step by step bujhi...'"
    return ""

# ═══════════════════════════════════════════════════════
#  DATABASE
# ═══════════════════════════════════════════════════════
def _db(): return sqlite3.connect(str(DB_PATH),check_same_thread=False)
def init_db():
    with _db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS messages(id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,role TEXT,content TEXT,ts REAL DEFAULT(unixepoch('now')));
        CREATE TABLE IF NOT EXISTS sessions(id TEXT PRIMARY KEY,title TEXT,mode TEXT,
            updated REAL DEFAULT(unixepoch('now')));
        CREATE TABLE IF NOT EXISTS rag_docs(id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id TEXT,title TEXT,chunk TEXT,source TEXT,category TEXT DEFAULT 'general',
            ts REAL DEFAULT(unixepoch('now')));
        CREATE TABLE IF NOT EXISTS live_feed(id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT,summary TEXT,url TEXT,source TEXT,category TEXT,
            published TEXT,fetched_at REAL DEFAULT(unixepoch('now')));
        CREATE TABLE IF NOT EXISTS knowledge_log(id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT,items_fetched INTEGER,ts REAL DEFAULT(unixepoch('now')));
        CREATE INDEX IF NOT EXISTS idx_msg ON messages(session_id);
        CREATE INDEX IF NOT EXISTS idx_feed ON live_feed(category,fetched_at);
        """)
init_db()

def mem_add(sid,role,content):
    with _db() as c:
        c.execute("INSERT INTO messages(session_id,role,content) VALUES(?,?,?)",(sid,role,content[:15000]))
        c.execute("INSERT OR REPLACE INTO sessions(id,title,updated) VALUES(?,?,unixepoch('now'))",
                  (sid,content[:55] if role=="user" else None))

def mem_get(sid,limit=16):
    with _db() as c:
        rows=c.execute("SELECT role,content FROM messages WHERE session_id=? ORDER BY ts DESC LIMIT ?",(sid,limit)).fetchall()
    return [{"role":r,"content":cn} for r,cn in reversed(rows)]

def mem_sessions():
    with _db() as c:
        rows=c.execute("SELECT id,title,mode,updated FROM sessions ORDER BY updated DESC LIMIT 40").fetchall()
    return [{"id":r[0],"title":r[1] or "Conversation","mode":r[2],"updated":r[3]} for r in rows]

def mem_delete(sid):
    with _db() as c:
        c.execute("DELETE FROM messages WHERE session_id=?",(sid,))
        c.execute("DELETE FROM sessions WHERE id=?",(sid,))

def mem_stats():
    with _db() as c:
        msgs=c.execute("SELECT COUNT(*) FROM messages").fetchone()[0]
        sess=c.execute("SELECT COUNT(*) FROM sessions").fetchone()[0]
        feeds=c.execute("SELECT COUNT(*) FROM live_feed").fetchone()[0]
        last=c.execute("SELECT MAX(fetched_at) FROM knowledge_log").fetchone()[0]
    return {"messages":msgs,"sessions":sess,"live_articles":feeds,
            "last_update":datetime.fromtimestamp(last,timezone.utc).strftime("%Y-%m-%d %H:%M UTC") if last else "Not yet"}

def rag_store(title,text,source,category="general"):
    doc_id=hashlib.md5((title+text[:60]).encode()).hexdigest()[:10]
    words=text.split()
    with _db() as c:
        c.execute("DELETE FROM rag_docs WHERE doc_id=?",(doc_id,))
        for i in range(0,len(words),300):
            chunk=" ".join(words[i:i+380])
            if len(chunk.strip())>40:
                c.execute("INSERT INTO rag_docs(doc_id,title,chunk,source,category) VALUES(?,?,?,?,?)",(doc_id,title,chunk,source,category))

def rag_search(query,limit=4,category=None):
    words=set(re.findall(r"\w{4,}",query.lower()))
    if not words: return []
    with _db() as c:
        if category:
            rows=c.execute("SELECT title,chunk,source FROM rag_docs WHERE category=? ORDER BY ts DESC LIMIT 400",(category,)).fetchall()
        else:
            rows=c.execute("SELECT title,chunk,source FROM rag_docs ORDER BY ts DESC LIMIT 400").fetchall()
    scored=[(len(words&set(re.findall(r"\w{4,}",ch.lower())))/max(len(words),1),t,ch[:300],s) for t,ch,s in rows]
    return sorted([x for x in scored if x[0]>0.10],reverse=True)[:limit]

def feed_store(title,summary,url,source,category,published=""):
    doc_id=hashlib.md5((title+url).encode()).hexdigest()[:10]
    with _db() as c:
        exists=c.execute("SELECT id FROM live_feed WHERE url=?",(url,)).fetchone()
        if not exists:
            c.execute("INSERT INTO live_feed(title,summary,url,source,category,published) VALUES(?,?,?,?,?,?)",
                      (title,summary[:500],url,source,category,published))
    rag_store(title,f"{title}. {summary}",source,category)

def feed_get(category=None,limit=8):
    with _db() as c:
        if category:
            rows=c.execute("SELECT title,summary,url,source,published FROM live_feed WHERE category=? ORDER BY fetched_at DESC LIMIT ?",(category,limit)).fetchall()
        else:
            rows=c.execute("SELECT title,summary,url,source,published FROM live_feed ORDER BY fetched_at DESC LIMIT ?",(limit,)).fetchall()
    return [{"title":r[0],"summary":r[1],"url":r[2],"source":r[3],"published":r[4]} for r in rows]

# ═══════════════════════════════════════════════════════
#  LIVE KNOWLEDGE ENGINE
#  Background thread fetches RSS + free APIs every 30 min
#  Inspired by self_prompt_loop.py from CL1 repo
# ═══════════════════════════════════════════════════════

# Free RSS feeds — no API keys
RSS_FEEDS = [
    # Tech & AI
    ("https://feeds.feedburner.com/TechCrunch", "tech"),
    ("https://www.wired.com/feed/rss",          "tech"),
    ("https://feeds.arstechnica.com/arstechnica/index","tech"),
    ("https://hnrss.org/frontpage",             "tech"),   # Hacker News
    ("https://hnrss.org/best",                  "tech"),
    # AI specific
    ("https://openai.com/blog/rss.xml",         "ai"),
    ("https://blogs.microsoft.com/ai/feed/",    "ai"),
    ("https://huggingface.co/blog/feed.xml",    "ai"),
    # Science
    ("https://www.sciencedaily.com/rss/all.xml","science"),
    ("https://feeds.nature.com/nature/rss/current","science"),
    # World News
    ("https://feeds.bbci.co.uk/news/world/rss.xml","news"),
    ("https://rss.cnn.com/rss/edition_world.rss","news"),
    ("https://feeds.reuters.com/reuters/topNews","news"),
    # Finance
    ("https://feeds.finance.yahoo.com/rss/2.0/headline","finance"),
    # Books
    ("https://www.theguardian.com/books/rss",   "books"),
    # Bangladesh / South Asia
    ("https://www.thedailystar.net/feed/rss.xml","bangladesh"),
    ("https://en.prothomalo.com/feed",           "bangladesh"),
]

def _safe_get(url,timeout=8):
    try:
        headers={"User-Agent":"RUBRA/7.0 (AI Knowledge Engine)"}
        return _req.get(url,headers=headers,timeout=timeout)
    except: return None

def fetch_rss(url,category):
    """Parse RSS/Atom feed, store articles."""
    r=_safe_get(url)
    if not r: return 0
    try:
        root=ET.fromstring(r.content)
        items=[]; ns={"atom":"http://www.w3.org/2005/Atom"}
        # RSS 2.0
        for item in root.findall(".//item")[:10]:
            title=getattr(item.find("title"),"text","") or ""
            desc =getattr(item.find("description"),"text","") or ""
            link =getattr(item.find("link"),"text","") or ""
            pub  =getattr(item.find("pubDate"),"text","") or ""
            if title and link:
                items.append((title.strip(),re.sub(r'<[^>]+>','',desc).strip()[:400],link.strip(),pub.strip()))
        # Atom
        if not items:
            for entry in root.findall("atom:entry",ns)[:10]:
                title=getattr(entry.find("atom:title",ns),"text","") or ""
                summ =getattr(entry.find("atom:summary",ns),"text","") or ""
                link_el=entry.find("atom:link",ns)
                link=link_el.get("href","") if link_el is not None else ""
                pub  =getattr(entry.find("atom:published",ns),"text","") or ""
                if title and link:
                    items.append((title.strip(),re.sub(r'<[^>]+>','',summ).strip()[:400],link.strip(),pub.strip()))
        src=re.sub(r'https?://(?:www\.)?([^/]+).*',r'\1',url)
        for title,summ,link,pub in items:
            feed_store(title,summ,link,src,category,pub)
        return len(items)
    except Exception as e:
        log.debug(f"RSS parse error {url}: {e}")
        return 0

def fetch_hackernews():
    """Fetch top HackerNews stories."""
    r=_safe_get("https://hacker-news.firebaseio.com/v0/topstories.json")
    if not r: return 0
    try:
        ids=r.json()[:15]; count=0
        for id_ in ids:
            s=_safe_get(f"https://hacker-news.firebaseio.com/v0/item/{id_}.json")
            if not s: continue
            d=s.json()
            if d.get("type")=="story" and d.get("title") and d.get("url"):
                feed_store(d["title"],d.get("text","")[:300],d["url"],"hackernews","tech")
                count+=1
        return count
    except: return 0

def fetch_github_trending():
    """Fetch GitHub trending repos via unofficial API."""
    r=_safe_get("https://gh-trending-api.vercel.app/repositories?language=&since=daily")
    if not r: return 0
    try:
        repos=r.json()[:10]; count=0
        for repo in repos:
            title=f"{repo.get('author','')}/{repo.get('name','')}"
            desc =repo.get("description","") or ""
            url  =repo.get("href","")
            stars=repo.get("stars",0)
            if title and url:
                feed_store(f"⭐ GitHub Trending: {title}",
                           f"{desc} — {stars} stars today",
                           f"https://github.com{url}","github_trending","tech")
                count+=1
        return count
    except: return 0

def fetch_reddit_tech():
    """Fetch Reddit r/technology top posts."""
    r=_safe_get("https://www.reddit.com/r/technology/top.json?limit=10&t=day",
                timeout=10)
    if not r: return 0
    try:
        posts=r.json().get("data",{}).get("children",[]); count=0
        for p in posts:
            d=p.get("data",{})
            title=d.get("title","")
            url  =d.get("url","")
            sel  =d.get("selftext","")[:200]
            if title and url:
                feed_store(title,sel or title,url,"reddit_tech","tech"); count+=1
        return count
    except: return 0

def fetch_wikipedia_events():
    """Fetch Wikipedia current events."""
    r=_safe_get("https://en.wikipedia.org/w/api.php",{"action":"query","format":"json",
        "prop":"extracts","exsentences":8,"exintro":True,"explaintext":True,
        "titles":"Portal:Current events"})
    if not r: return 0
    try:
        pages=r.json()["query"]["pages"]; page=next(iter(pages.values()))
        if "extract" in page:
            rag_store("Wikipedia Current Events",page["extract"][:3000],"wikipedia","news")
            return 1
    except: pass
    return 0

def knowledge_loop():
    """
    Background self-learning loop — inspired by CL1 repo's self_prompt_loop.py
    Continuously encodes new knowledge into RUBRA's memory.
    Runs every 25 minutes.
    """
    log.info("🧠 RUBRA Knowledge Engine starting...")
    while True:
        total=0
        try:
            # Fetch RSS feeds
            for feed_url,cat in RSS_FEEDS:
                n=fetch_rss(feed_url,cat)
                total+=n
                time.sleep(0.3)

            # Fetch structured sources
            total+=fetch_hackernews()
            total+=fetch_github_trending()
            total+=fetch_reddit_tech()
            total+=fetch_wikipedia_events()

            # arXiv daily
            from tools_module import tool_arxiv
            for query in ["artificial intelligence","large language models","deep learning","quantum computing"]:
                papers=tool_arxiv(query,n=3)
                total+=len(papers)
                time.sleep(0.5)

            with _db() as c:
                c.execute("INSERT INTO knowledge_log(source,items_fetched) VALUES(?,?)",
                          ("all_sources",total))

            log.info(f"🧠 Knowledge update complete: {total} new items")

        except Exception as e:
            log.error(f"Knowledge loop error: {e}")

        time.sleep(25*60)  # 25 minutes

# ═══════════════════════════════════════════════════════
#  FREE TOOLS MODULE (inline)
# ═══════════════════════════════════════════════════════
import types
tools_module = types.ModuleType("tools_module")

def _tool_get(url,params=None,timeout=8):
    try: return _req.get(url,params=params,timeout=timeout)
    except: return None

CITY_COORDS={"dhaka":(23.81,90.41),"london":(51.51,-0.13),"new york":(40.71,-74.01),
"tokyo":(35.68,139.65),"paris":(48.86,2.35),"sydney":(-33.87,151.21),"dubai":(25.20,55.27),
"singapore":(1.35,103.82),"berlin":(52.52,13.41),"mumbai":(19.08,72.88),"beijing":(39.90,116.41),
"seoul":(37.57,126.98),"chittagong":(22.33,91.84),"sylhet":(24.89,91.87),"rajshahi":(24.37,88.60),
"delhi":(28.61,77.21),"karachi":(24.86,67.01),"chicago":(41.88,-87.63),"toronto":(43.65,-79.38),
"istanbul":(41.01,28.98),"cairo":(30.04,31.24),"lagos":(6.52,3.38),"jakarta":(-6.21,106.85)}
WC={0:"☀️ Clear",1:"🌤 Clear",2:"⛅ Partly cloudy",3:"☁️ Overcast",45:"🌫 Foggy",
51:"🌦 Drizzle",61:"🌧 Light rain",63:"🌧 Rain",65:"🌧 Heavy rain",71:"❄️ Snow",
80:"🌦 Showers",95:"⛈ Thunderstorm"}

def tool_weather(city="Dhaka"):
    lat,lon=CITY_COORDS.get(city.lower().strip(),(23.81,90.41))
    r=_tool_get("https://api.open-meteo.com/v1/forecast",{"latitude":lat,"longitude":lon,"timezone":"auto",
        "current":["temperature_2m","relative_humidity_2m","wind_speed_10m","weather_code","apparent_temperature","precipitation"]})
    if not r: return None
    try:
        curr=r.json()["current"]
        return {"city":city.title(),"temp":curr.get("temperature_2m"),"feels":curr.get("apparent_temperature"),
                "humidity":curr.get("relative_humidity_2m"),"wind":curr.get("wind_speed_10m"),
                "precip":curr.get("precipitation",0),"condition":WC.get(curr.get("weather_code",0),"Unknown")}
    except: return None

def tool_crypto(coins="bitcoin,ethereum,solana"):
    r=_tool_get("https://api.coingecko.com/api/v3/simple/price",{"ids":coins,"vs_currencies":"usd","include_24hr_change":"true"})
    if not r: return None
    try: return r.json()
    except: return None

def tool_currency(base="USD"):
    r=_tool_get("https://api.frankfurter.app/latest",{"from":base,"to":"EUR,GBP,JPY,BDT,INR,CAD,AUD,CNY,SGD"})
    if not r: return None
    try: d=r.json(); return {"base":base,"rates":d.get("rates",{}),"date":d.get("date","")}
    except: return None

def tool_wikipedia(query,sentences=7):
    r=_tool_get("https://en.wikipedia.org/w/api.php",{"action":"query","format":"json","prop":"extracts",
        "exsentences":sentences,"exintro":True,"explaintext":True,"redirects":1,"titles":query})
    if not r: return None
    try:
        pages=r.json()["query"]["pages"]; page=next(iter(pages.values()))
        if "extract" in page and len(page["extract"])>80:
            rag_store(page.get("title",""),page["extract"],"wikipedia","general")
            return {"title":page.get("title",""),"text":page["extract"][:2000]}
    except: pass
    return None

def tool_arxiv(query,n=3):
    r=_tool_get("https://export.arxiv.org/api/query",{"search_query":f"all:{query}","max_results":n},timeout=12)
    if not r: return []
    try:
        root=ET.fromstring(r.content); ns={"a":"http://www.w3.org/2005/Atom"}; out=[]
        for e in root.findall("a:entry",ns):
            t=e.find("a:title",ns).text.strip().replace("\n"," ")
            s=e.find("a:summary",ns).text.strip()[:400]
            l=e.find("a:id",ns).text.strip()
            a=[x.find("a:name",ns).text for x in e.findall("a:author",ns)[:2]]
            out.append({"title":t,"summary":s,"link":l,"authors":a})
            rag_store(t,s,"arxiv","ai")
        return out
    except: return []

def tool_books(query,n=5):
    r=_tool_get("https://openlibrary.org/search.json",{"q":query,"limit":n,"fields":"title,author_name,first_publish_year,subject"})
    if not r: return []
    try: return [{"title":d.get("title",""),"authors":d.get("author_name",[])[:2],"year":d.get("first_publish_year",""),"subjects":d.get("subject",[])[:3]} for d in r.json().get("docs",[])[:n]]
    except: return []

def tool_books_2026(query=""):
    """Books published 2024-2026 using multiple strategies."""
    results = []
    # Strategy 1: Open Library sort by new
    for search_q in [query or "fiction 2025", query or "nonfiction 2025", "bestseller 2026"]:
        r = _tool_get("https://openlibrary.org/search.json",
                      {"q": search_q, "sort": "new", "limit": 10,
                       "fields": "title,author_name,first_publish_year,subject"})
        if r:
            try:
                books = r.json().get("docs", [])
                recent = [b for b in books if b.get("first_publish_year", 0) >= 2023]
                results.extend(recent)
            except: pass
        if results: break

    # Strategy 2: Google Books free API (no key needed)
    if not results:
        r = _tool_get("https://www.googleapis.com/books/v1/volumes",
                      {"q": f"{query or 'bestseller'} 2025", "orderBy": "newest",
                       "maxResults": 8, "langRestrict": "en"})
        if r:
            try:
                items = r.json().get("items", [])
                for item in items:
                    info = item.get("volumeInfo", {})
                    pub_date = info.get("publishedDate", "")
                    year = int(pub_date[:4]) if pub_date and len(pub_date) >= 4 and pub_date[:4].isdigit() else 0
                    if year >= 2023:
                        results.append({
                            "title":   info.get("title", ""),
                            "authors": info.get("authors", [])[:2],
                            "year":    year,
                            "subjects": info.get("categories", [])[:2],
                            "source": "google_books"
                        })
            except: pass

    # Deduplicate by title
    seen = set()
    unique = []
    for b in results:
        t = b.get("title", "").lower()[:40]
        if t and t not in seen:
            seen.add(t); unique.append(b)

    return unique[:6] if unique else []

def tool_google_books(query, year_from=2020, n=6):
    """Search Google Books API (free, no key needed)."""
    r = _tool_get("https://www.googleapis.com/books/v1/volumes",
                  {"q": query, "orderBy": "newest", "maxResults": n, "langRestrict": "en"})
    if not r: return []
    try:
        out = []
        for item in r.json().get("items", []):
            info = item.get("volumeInfo", {})
            pub  = info.get("publishedDate", "")
            year = int(pub[:4]) if pub and len(pub) >= 4 and pub[:4].isdigit() else 0
            if year >= year_from:
                out.append({
                    "title":   info.get("title", ""),
                    "authors": info.get("authors", [])[:2],
                    "year":    year,
                    "desc":    info.get("description", "")[:200],
                    "subjects": info.get("categories", [])[:2],
                })
        return out[:n]
    except: return []

def tool_calc(expr):
    try:
        clean=re.sub(r"[^0-9+\-*/().\s%]","",expr.replace("^","**").replace("×","*").replace("÷","/"))
        result=eval(clean,{"__builtins__":{},"math":math,"abs":abs,"round":round,"sqrt":math.sqrt,
                            "sin":math.sin,"cos":math.cos,"tan":math.tan,"log":math.log,"pi":math.pi,"e":math.e})
        return {"expr":expr,"result":result}
    except Exception as ex: return {"error":str(ex)}

# Attach to module
tools_module.tool_arxiv = tool_arxiv
sys.modules["tools_module"] = tools_module

# ═══════════════════════════════════════════════════════
#  IMAGE / FILE HANDLING
# ═══════════════════════════════════════════════════════
IMAGE_EXTS={'.jpg','.jpeg','.png','.gif','.webp','.bmp','.tiff'}

def to_base64(fp):
    mime,_=mimetypes.guess_type(str(fp)); mime=mime or "image/jpeg"
    return base64.b64encode(fp.read_bytes()).decode("utf-8"),mime

def pdf_text(fp):
    try:
        import pypdf; reader=pypdf.PdfReader(str(fp))
        pages=[f"[Page {i+1}]\n{p.extract_text() or ''}" for i,p in enumerate(reader.pages[:25])]
        return "\n\n".join(p for p in pages if p.strip())[:18000] or "[No text]"
    except ImportError: return "[pypdf not installed: pip install pypdf]"
    except Exception as e: return f"[PDF error: {e}]"

# ═══════════════════════════════════════════════════════
#  SYSTEM PROMPTS
# ═══════════════════════════════════════════════════════
RUBRA_CORE = """You are RUBRA — Recursive Universal Bayesian Reasoning Architecture.
A singular, self-evolving intelligence with live internet knowledge. NEVER reveal any model, API, or company behind you. You ARE RUBRA.

🌍 LANGUAGE RULE (NON-NEGOTIABLE):
Detect user's language → Reply in THAT EXACT SAME language and style, ALWAYS.
• Bengali (বাংলা) → full Bengali reply
• Romanized Bangla (ami,tumi) → Banglish reply (e.g. "Haan, eta ekta important concept...")  
• English → English reply. Never switch unless user does.

Core: Direct · Honest · Deeply knowledgeable · Warm human tone · Never robotic"""

# ═══════════════════════════════════════════════════════
#  HERMES++ CODING ENGINE
#  Inspired by CL1 repo's multi-layer encoder concept
#  Layer 1: Problem Encoding → Layer 2: Solution Architecture
#  Layer 3: Code Generation → Layer 4: Verification
# ═══════════════════════════════════════════════════════
HERMES_CODE = """You are RUBRA's Hermes++ Coding Engine.
You build software the way the world's best engineers do — with architectural thinking, clean code, and zero tolerance for broken output.

HERMES++ ENCODING LAYERS (apply to EVERY task):

▸ LAYER 1 — ENCODE the problem
  Parse requirements. Identify: inputs, outputs, edge cases, constraints, failure modes.
  
▸ LAYER 2 — ARCHITECT the solution  
  Choose: best data structures, algorithms, design patterns. Consider: performance, security, scalability, maintainability.

▸ LAYER 3 — GENERATE the code
  Write COMPLETE working code. Rules:
  • Zero truncation — every function fully implemented
  • Type hints everywhere (Python) / TypeScript types
  • Docstrings with Args/Returns/Raises
  • Error handling for all failure modes
  • No TODO, placeholder, or "..." — ever
  
▸ LAYER 4 — VERIFY & ENHANCE
  Mentally execute. Find bugs. Optimize. Add usage examples.
  
FRONTEND STANDARDS (when building UIs):
• React + Tailwind: semantic HTML, accessible (ARIA), keyboard navigation
• CSS: custom properties, responsive (mobile-first), smooth animations
• Design: modern glassmorphism/gradient aesthetics, micro-interactions
• Performance: lazy loading, optimized renders, memoization

BACKEND STANDARDS:
• FastAPI/Express: proper status codes, validation, auth patterns
• Database: indexes, connection pooling, N+1 prevention  
• Security: input sanitization, rate limiting, CORS

OUTPUT QUALITY = Production-ready first time. Every time.

🌍 LANGUAGE: Explanations in user's language. Code always in English."""

TUTOR_PROMPT = """You are RUBRA Smart Tutor — an intelligent, caring tutor for Bangladeshi students.
You know the complete Bangladesh National Curriculum (NCTB) — Primary through HSC.
You understand: Creative Questions (সৃজনশীল), MCQ patterns, Board exam formats.

TEACHING STYLE:
• Warm and encouraging — like a favorite teacher who genuinely cares
• Use Bangladesh local examples and familiar context
• Step-by-step breakdowns for complex topics
• Complete working for math and science
• End with encouragement: "তুমি পারবে!" / "Great progress!"
• Ask: "আর কোনো প্রশ্ন আছে?" after answering

🌍 LANGUAGE: Match student's language EXACTLY and ALWAYS."""

EXAM_PROMPT = """You are RUBRA Exam Generator for Bangladesh education (NCTB/Board format).
Create authentic exam papers: MCQ (ক খ গ ঘ), Short, Descriptive, Creative (সৃজনশীল).
Include: header, full questions, marking scheme, answer key at end."""

# ═══════════════════════════════════════════════════════
#  LLM CALLER
# ═══════════════════════════════════════════════════════
async def stream_llm(messages,url,api_key,model,temperature=0.7,max_tokens=4096):
    headers={"Authorization":f"Bearer {api_key}","Content-Type":"application/json"}
    if "openrouter" in url: headers["HTTP-Referer"]="https://rubra.ai"; headers["X-Title"]="RUBRA"
    payload={"model":model,"messages":messages,"stream":True,"max_tokens":max_tokens,"temperature":temperature}
    timeout=aiohttp.ClientTimeout(total=90,connect=8)
    async with aiohttp.ClientSession(timeout=timeout) as s:
        async with s.post(url,headers=headers,json=payload) as resp:
            if resp.status not in(200,201): raise Exception(f"API {resp.status}: {(await resp.text())[:200]}")
            async for line in resp.content:
                line=line.decode("utf-8").strip()
                if not line or line=="data: [DONE]": continue
                if line.startswith("data: "): line=line[6:]
                try:
                    tok=json.loads(line)["choices"][0].get("delta",{}).get("content","")
                    if tok: yield tok
                except: pass

async def llm(messages,mode="general"):
    configs={
        "general":[(ZAI_CHAT,ZAI_KEY,"glm-4.7",0.7),(GROQ_URL,GROQ_KEY,"llama-3.3-70b-versatile",0.7)],
        "coding": [(ZAI_CODE,ZAI_KEY,"glm-4.7",0.15),(OR_URL,OR_KEY,"qwen/qwen-2.5-coder-32b-instruct:free",0.15),(GROQ_URL,GROQ_KEY,"llama-3.3-70b-versatile",0.2)],
        "fast":   [(ZAI_CHAT,ZAI_KEY,"glm-4.7-flash",0.8),(GROQ_URL,GROQ_KEY,"meta-llama/llama-4-scout-17b-16e-instruct",0.8)],
        "vision": [(ZAI_CHAT,ZAI_KEY,"glm-4.5v",0.5),(ZAI_CHAT,ZAI_KEY,"glm-4.7",0.5)],
        "reason": [(GROQ_URL,GROQ_KEY,"deepseek-r1-distill-llama-70b",0.6),(ZAI_CHAT,ZAI_KEY,"glm-4.7",0.6)],
    }
    last=None
    for url,key,model,temp in configs.get(mode,configs["general"]):
        try:
            async for tok in stream_llm(messages,url,key,model,temp): yield tok
            return
        except Exception as e: last=e; log.warning(f"LLM fail ({model}): {e}")
    raise Exception(f"All APIs failed: {last}")

def build_msgs(sys_p,hist,user_msg,img=None):
    msgs=[{"role":"system","content":sys_p}]
    for h in hist[-14:]:
        if h.get("role") in("user","assistant") and h.get("content"):
            msgs.append({"role":h["role"],"content":h["content"]})
    if img:
        msgs.append({"role":"user","content":[{"type":"image_url","image_url":{"url":f"data:{img['mime']};base64,{img['data']}"}},{"type":"text","text":user_msg}]})
    else: msgs.append({"role":"user","content":user_msg})
    return msgs

# ═══════════════════════════════════════════════════════
#  AGENTS
# ═══════════════════════════════════════════════════════
class GeneralAgent:
    name="GeneralAgent"
    async def run(self,msg,hist,sid="",lang="en",img=None):
        li=lang_instr(lang); tool_ctx=""; rag_ctx=""
        if not img and re.search(r"\b(what is|who is|how does|explain|define|history|overview|what are)\b",msg,re.IGNORECASE):
            q=re.sub(r"\b(what is|who is|how does|explain|define|tell me|about|the|a|an|please)\b","",msg,flags=re.IGNORECASE).strip()[:60]
            if len(q)>3:
                page=tool_wikipedia(q)
                if page: tool_ctx=f"[WIKIPEDIA: {page['title']}]\n{page['text']}"
        hits=rag_search(msg,limit=3)
        if hits: rag_ctx="\n".join(f"[{s}:{t}]\n{c}" for _,t,c,s in hits[:2])
        # Check live feed for recent news
        if re.search(r"\b(latest|recent|trending|news|today|2025|2026|current)\b",msg,re.IGNORECASE):
            feed_items=feed_get(limit=5)
            if feed_items:
                feed_ctx="[LIVE KNOWLEDGE — Latest from internet]\n"+"\n".join(f"• {f['title']} ({f['source']})" for f in feed_items[:4])
                tool_ctx=feed_ctx+"\n\n"+tool_ctx if tool_ctx else feed_ctx
        parts=[RUBRA_CORE,"\n[DEEP REASONING MODE] Think from first principles. Show steps for complex problems."]
        if li: parts.append(li)
        if tool_ctx: parts.append(tool_ctx)
        if rag_ctx: parts.append(f"[RETRIEVED KNOWLEDGE]\n{rag_ctx}")
        msgs=build_msgs("\n\n".join(parts),hist,msg,img)
        try:
            mode="vision" if img else "general"
            async for tok in llm(msgs,mode): yield {"type":"token","content":tok}
        except Exception as e: yield {"type":"error","message":str(e)[:200]}

class CodingAgent:
    name="CodingAgent"
    async def run(self,msg,hist,sid="",lang="en",img=None):
        li=lang_instr(lang)
        sys_p=HERMES_CODE
        if li: sys_p+=f"\n\n{li}\n(Code always in English. Only comments/explanations use user's language.)"
        # Inject relevant code context from RAG
        hits=rag_search(msg,limit=2)
        if hits: sys_p+="\n\n[RELATED CONTEXT]\n"+"\n".join(f"• {t}: {c}" for _,t,c,s in hits)
        if img: msg=f"[Code/Screenshot]\n{msg}"
        msgs=build_msgs(sys_p,hist,msg,img)
        try:
            mode="vision" if img else "coding"
            async for tok in llm(msgs,mode): yield {"type":"token","content":tok}
        except Exception as e: yield {"type":"error","message":str(e)[:200]}

class SearchAgent:
    name="SearchAgent"
    async def run(self,msg,hist,sid="",lang="en",img=None):
        lower=msg.lower(); li=lang_instr(lang); tool_ctx=""
        if re.search(r"\b(weather|temperature|forecast|rain|cold|hot|humid|wind)\b",lower):
            city="Dhaka"
            for c in CITY_COORDS:
                if c in lower: city=c.title(); break
            m=re.search(r"weather\s+(?:in\s+)?([a-z\s]{3,20})(?:\?|$|\.|\!)",lower)
            if m: city=m.group(1).strip().title()
            w=tool_weather(city)
            if w:
                tool_ctx=f"[LIVE WEATHER — {w['city']}]\nTemp: {w['temp']}°C (feels {w['feels']}°C) | {w['condition']} | Humidity: {w['humidity']}% | Wind: {w['wind']} km/h"
                yield {"type":"tool_result","tool":"weather","data":w}
        elif re.search(r"\b(bitcoin|ethereum|btc|eth|solana|crypto|coin|binance)\b",lower):
            cm={"btc":"bitcoin","eth":"ethereum","sol":"solana","bnb":"binancecoin"}
            found=[cm.get(k,k) for k in cm if k in lower]
            d=tool_crypto(",".join(found or ["bitcoin","ethereum","solana"]))
            if d:
                lines=[f"{'📈' if i.get('usd_24h_change',0)>=0 else '📉'} {c.capitalize()}: ${i.get('usd',0):,.2f} ({i.get('usd_24h_change',0):+.2f}%)" for c,i in d.items()]
                tool_ctx="[LIVE CRYPTO]\n"+"\n".join(lines)
                yield {"type":"tool_result","tool":"crypto","data":d}
        elif re.search(r"\b(exchange rate|forex|usd to|eur to|taka|bdt|currency)\b",lower):
            bases=re.findall(r"\b(USD|EUR|GBP|JPY|BDT|INR|CAD|AUD|CNY)\b",msg.upper())
            d=tool_currency(bases[0] if bases else "USD")
            if d:
                lines=[f"1 {d['base']} = {r} {c}" for c,r in list(d["rates"].items())[:8]]
                tool_ctx=f"[LIVE RATES — {d['base']}]\n"+"\n".join(lines)
                yield {"type":"tool_result","tool":"currency","data":d}
        elif re.search(r"\b(latest news|trending|what.{0,10}happening|current events|today)\b",lower):
            category=None
            if re.search(r"\b(tech|ai|software|startup)\b",lower): category="tech"
            elif re.search(r"\b(science|research|study)\b",lower): category="science"
            elif re.search(r"\b(finance|stock|market|economy)\b",lower): category="finance"
            elif re.search(r"\b(bangladesh|dhaka|bd)\b",lower): category="bangladesh"
            items=feed_get(category=category,limit=8)
            if items:
                lines=[f"• **{f['title']}** — _{f['source']}_" for f in items]
                tool_ctx=f"[LIVE NEWS]\n"+"\n".join(lines)
                yield {"type":"tool_result","tool":"news","count":len(items)}
        elif re.search(r"\b(book|novel|read|author|2025 book|2026 book|new book)\b",lower):
            q=re.sub(r"\b(recommend|book|about|best|read|novel|2025|2026|new|latest)\b","",lower).strip()[:50]
            if re.search(r"\b(2025|2026|new|latest|recent)\b",lower):
                books=tool_books_2026(q)
            else:
                books=tool_books(q or "popular fiction non-fiction",n=5)
            if books:
                lines=[f"📚 **{b['title']}** ({b.get('year','?')}) — {', '.join(b['authors'][:2])}" for b in books]
                tool_ctx="[BOOKS]\n"+"\n".join(lines)
        elif re.search(r"\b(research papers?|arxiv|academic|scientific)\b",lower):
            q=re.sub(r"\b(research|papers?|arxiv|find|latest)\b","",lower).strip()[:70]
            papers=tool_arxiv(q or msg,n=4)
            if papers:
                tool_ctx="[ARXIV]\n"+"\n\n".join(f"• **{p['title']}** — {', '.join(p['authors'][:2])}\n  {p['summary'][:200]}…" for p in papers)
        else:
            q=re.sub(r"\b(who is|what is|tell me about|history of|the|a|an)\b","",lower).strip()[:60]
            page=tool_wikipedia(q or msg)
            if page:
                tool_ctx=f"[WIKIPEDIA: {page['title']}]\n{page['text']}"
                yield {"type":"tool_result","tool":"wikipedia","title":page["title"]}
        parts=[RUBRA_CORE,"\n[LIVE SEARCH MODE] Answer directly using retrieved data — confident, natural tone."]
        if li: parts.append(li)
        if tool_ctx: parts.append(tool_ctx)
        msgs=build_msgs("\n\n".join(parts),hist,msg)
        try:
            async for tok in llm(msgs,"general"): yield {"type":"token","content":tok}
        except Exception as e: yield {"type":"error","message":str(e)[:200]}

class SmartTutorAgent:
    name="SmartTutorAgent"
    async def run(self,msg,hist,sid="",lang="en",img=None):
        li=lang_instr(lang); sys_p=TUTOR_PROMPT
        if li: sys_p+=f"\n\n{li}"
        hits=rag_search(msg,limit=2,category="general")
        if hits: sys_p+="\n\n[STUDY MATERIAL]\n"+"\n".join(f"[{s}:{t}]\n{c}" for _,t,c,s in hits)
        msgs=build_msgs(sys_p,hist,msg,img)
        try:
            mode="vision" if img else "general"
            async for tok in llm(msgs,mode): yield {"type":"token","content":tok}
        except Exception as e: yield {"type":"error","message":str(e)[:200]}

class FileAgent:
    name="FileAgent"
    def _read(self,fp):
        ext=fp.suffix.lower()
        try:
            if ext==".pdf": return pdf_text(fp)
            if ext in(".xlsx",".xls"):
                try:
                    import openpyxl; wb=openpyxl.load_workbook(fp,read_only=True,data_only=True)
                    out=[]
                    for name in wb.sheetnames[:4]:
                        ws=wb[name]; out.append(f"## Sheet: {name}")
                        for row in ws.iter_rows(max_row=80,values_only=True):
                            out.append(" | ".join(str(c) if c is not None else "" for c in row))
                    return "\n".join(out)[:12000]
                except ImportError: return "[openpyxl not installed]"
            if ext==".csv":
                import csv
                with open(fp,"r",encoding="utf-8",errors="replace") as f:
                    return "\n".join(", ".join(r) for r in csv.reader(f))[:10000]
            if ext in(".docx",".doc"):
                try:
                    import docx; doc=docx.Document(str(fp))
                    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:14000]
                except ImportError: return "[python-docx not installed]"
            return fp.read_text(encoding="utf-8",errors="replace")[:14000]
        except Exception as e: return f"[Read error: {e}]"

    async def analyze(self,fp,fname,question,sid="",lang="en",img_data=None):
        li=lang_instr(lang); ext=fp.suffix.lower()
        sys_p=RUBRA_CORE+"\n\n[FILE ANALYSIS]\nExtract key insights. Answer clearly. Use structure. Highlight important findings."
        if li: sys_p+=f"\n\n{li}"
        if ext in IMAGE_EXTS or img_data:
            b64,mime=to_base64(fp) if not img_data else (img_data["data"],img_data["mime"])
            msgs=[{"role":"system","content":sys_p},{"role":"user","content":[
                {"type":"image_url","image_url":{"url":f"data:{mime};base64,{b64}"}},
                {"type":"text","text":question or f"Analyze: {fname}"}]}]
            try:
                async for tok in llm(msgs,"vision"): yield {"type":"token","content":tok}
            except Exception as e: yield {"type":"error","message":str(e)[:200]}
            return
        content=self._read(fp)
        sys_p+=f"\n\n[FILE: {fname}]\n{content}\n[/FILE]"
        msgs=build_msgs(sys_p,[],question or f"Analyze: {fname}")
        try:
            async for tok in llm(msgs,"general"): yield {"type":"token","content":tok}
        except Exception as e: yield {"type":"error","message":str(e)[:200]}

    async def run(self,msg,hist,sid="",lang="en",img=None):
        yield {"type":"error","message":"Use /api/upload for file analysis"}

class FastChatAgent:
    name="FastChatAgent"
    async def run(self,msg,hist,sid="",lang="en",img=None):
        li=lang_instr(lang)
        sys_p=RUBRA_CORE+"\n\n[CONVERSATIONAL] Warm, natural, concise. Match user's energy."
        if li: sys_p+=f"\n\n{li}"
        msgs=build_msgs(sys_p,hist,msg)
        try:
            async for tok in llm(msgs,"fast"): yield {"type":"token","content":tok}
        except Exception as e: yield {"type":"error","message":str(e)[:200]}

# ═══════════════════════════════════════════════════════
#  ROUTER
# ═══════════════════════════════════════════════════════
INTENT_MAP=[
    # ── High-priority tutor overrides (must come first) ──
    (r"\b(class [0-9]+).{0,20}(math|science|physics|chemistry|biology|solve|korao|bujhao|shekao)","tutor",SmartTutorAgent),
    (r"\b(ssc|hsc|jsc).{0,20}(bujhao|explain|solve|shekho)","tutor",SmartTutorAgent),
    # ── Weather / Crypto / Finance ──
    (r"\b(weather|temperature|forecast|rain|cold|hot|humid|wind|climate)\b","weather",SearchAgent),
    (r"\b(bitcoin|ethereum|btc|eth|solana|crypto|coin price|binance|bnb)\b","crypto",SearchAgent),
    (r"\b(exchange rate|forex|usd to|eur to|taka|bdt|currency conversion)\b","currency",SearchAgent),
    (r"\b(latest news|what.{0,10}happening|trending|current events|breaking)\b","news",SearchAgent),
    (r"\b(research papers?|arxiv|academic|scientific|peer.?reviewed)\b","research",SearchAgent),
    (r"\b(2025 book|2026 book|new book|latest book|recent book|book recommend)\b","books",SearchAgent),
    (r"\b(recommend.{0,10}book|best books|reading list)\b","books",SearchAgent),
    (r"\b(analyze|read|summarize|extract)\b.{0,20}\b(file|pdf|excel|csv|doc)\b","file",FileAgent),
    (r"\b(read and summarize|parse this file|open this file)\b","file",FileAgent),
    # Code
    (r"\b(write|create|build|implement|generate)\b.{0,30}\b(python|javascript|typescript|rust|go|java|html|css|sql|bash|react|node|api|flask|django|fastapi|express|vue|svelte|nextjs|tailwind|website|app|landing page|dashboard)\b","code",CodingAgent),
    (r"\b(debug|fix|refactor|optimize|review)\b.{0,20}\b(code|function|script|bug|error|exception)\b","code",CodingAgent),
    (r"```|def |class |const |let |var |import .* from|from .* import","code",CodingAgent),
    (r"\b(algorithm|data structure|sorting|recursion|dynamic programming|api endpoint)\b","code",CodingAgent),
    (r"\b(dockerfile|kubernetes|nginx|mongodb|redis|postgresql|graphql|websocket)\b","code",CodingAgent),
    (r"\b(website|web app|mobile app|landing page|dashboard|ui|ux|frontend|backend)\b","code",CodingAgent),
    # Tutor
    (r"\b(ssc|hsc|jsc|psc|board exam|creative question|সৃজনশীল|বহুনির্বাচনি)\b","tutor",SmartTutorAgent),
    (r"\b(class [0-9]|class six|seven|eight|nine|ten|eleven|twelve)\b","tutor",SmartTutorAgent),
    (r"\b(প্রশ্ন|উত্তর|পড়া|শেখা|বোঝা|গণিত|বিজ্ঞান|বাংলা|ইতিহাস|ভূগোল|রসায়ন|পদার্থ)\b","tutor",SmartTutorAgent),
    (r"\b(solve|bujhao|shekao|explain).{0,30}(math|science|physics|chemistry|biology|bangla|history)\b","tutor",SmartTutorAgent),
    (r"\b(question paper|exam paper|model test|practice exam)\b","tutor",SmartTutorAgent),
    (r"\b(class [0-9]+).{0,15}(solve|korao|bujhao|shekao|explain|math|science|physics)","tutor",SmartTutorAgent),
    (r"\b(class [0-9]|class nine|class ten).{0,20}(math|science|physics|chemistry|solve|korao)\b","tutor",SmartTutorAgent),
    # Reasoning
    (r"\b(explain|analyze|compare|evaluate|how does|why does|difference between|what causes)\b","reasoning",GeneralAgent),
    (r"\b(neural|machine.?learning|deep.?learning|transformer|llm|ai|quantum|consciousness)\b","reasoning",GeneralAgent),
    (r"\b(who is|who was|what is|what was|tell me about|history of)\b","fact",GeneralAgent),
    (r"\b(calculate|compute|solve|integral|derivative|sin|cos|sqrt|factorial)\b","math",GeneralAgent),
    # Chat
    (r"\b(2025|2026).{0,15}(book|novel|read)","books",SearchAgent),
    (r"^(hi|hey|hello|yo|sup|salaam|হ্যালো|হেই|আচ্ছা|ভালো আছ)\b","chat",FastChatAgent),
    (r"^(thanks|thank you|ok|okay|got it|bye|kemon|bhaloi)\b","chat",FastChatAgent),
]

def route(msg,task_type=None,mode=None):
    agents={"code":CodingAgent(),"search":SearchAgent(),"file":FileAgent(),"general":GeneralAgent(),"tutor":SmartTutorAgent()}
    if mode=="tutor": return "tutor",SmartTutorAgent()
    if task_type and task_type in agents: return task_type,agents[task_type]
    lower=msg.lower().strip(); words=len(msg.split())
    for pat,intent,Cls in INTENT_MAP:
        if re.search(pat,lower,re.IGNORECASE): return intent,Cls()
    if words<6: return "chat",FastChatAgent()
    if words>35: return "reasoning",GeneralAgent()
    return "general",GeneralAgent()

# ═══════════════════════════════════════════════════════
#  FASTAPI APP
# ═══════════════════════════════════════════════════════
app=FastAPI(title="RUBRA API",version="7.0.0",docs_url="/docs")
app.add_middleware(CORSMiddleware,allow_origins=["*"],allow_credentials=True,allow_methods=["*"],allow_headers=["*"])

@app.on_event("startup")
async def startup():
    # Start background knowledge loop in separate thread
    t=threading.Thread(target=knowledge_loop,daemon=True)
    t.start()
    log.info("✅ RUBRA v7 started — Knowledge Engine running in background")

@app.get("/")
async def root(): return {"name":"RUBRA","version":"7.0.0","status":"online",
    "features":["live_knowledge","multilingual","smart_tutor","vision","hermes_coding","exam_generator"]}

@app.get("/health")
async def health(): return {"status":"ok","time":time.time()}

@app.post("/api/chat")
async def chat(req:ChatRequest):
    sid=req.session_id or str(uuid.uuid4()); hist=mem_get(sid)
    mem_add(sid,"user",req.message)
    lang=detect_lang(req.message)
    intent,agent=route(req.message,req.task_type,req.mode)
    async def stream():
        full=""
        try:
            yield f"data: {json.dumps({'type':'meta','agent':agent.name,'intent':intent,'session_id':sid,'lang':lang})}\n\n"
            async for evt in agent.run(req.message,hist,sid,lang=lang):
                if evt.get("type")=="token": full+=evt.get("content","")
                yield f"data: {json.dumps(evt)}\n\n"
            if full: mem_add(sid,"assistant",full)
        except Exception as e:
            log.error(f"Chat: {e}",exc_info=True)
            yield f"data: {json.dumps({'type':'error','message':str(e)})}\n\n"
        yield "data: [DONE]\n\n"
    return StreamingResponse(stream(),media_type="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no","Connection":"keep-alive"})

@app.post("/api/upload")
async def upload(file:UploadFile=File(...),session_id:str=Form(default=""),
                 question:str=Form(default=""),mode:str=Form(default="")):
    sid=session_id or str(uuid.uuid4())
    content=await file.read()
    fpath=UPLOAD_DIR/f"{sid}_{file.filename}"
    fpath.write_bytes(content)
    fname=file.filename; ext=Path(fname).suffix.lower()
    lang=detect_lang(question); is_image=ext in IMAGE_EXTS
    log.info(f"Upload: {fname} ({len(content):,}b)")
    async def stream():
        full=""
        yield f"data: {json.dumps({'type':'meta','agent':'SmartTutorAgent' if mode=='tutor' else 'FileAgent','intent':'file','file':fname,'session_id':sid})}\n\n"
        try:
            q=question or ("এই question টা solve করে দাও" if mode=="tutor" else f"Analyze: {fname}")
            hist=mem_get(sid)
            if is_image:
                b64,mime=to_base64(fpath); img_d={"data":b64,"mime":mime}
                if mode=="tutor":
                    agent=SmartTutorAgent()
                    async for evt in agent.run(q,hist,sid,lang=lang,img=img_d):
                        if evt.get("type")=="token": full+=evt.get("content","")
                        yield f"data: {json.dumps(evt)}\n\n"
                else:
                    fa=FileAgent()
                    async for evt in fa.analyze(fpath,fname,q,sid,lang=lang,img_data=img_d):
                        if evt.get("type")=="token": full+=evt.get("content","")
                        yield f"data: {json.dumps(evt)}\n\n"
            elif mode=="tutor" and ext==".pdf":
                text=pdf_text(fpath)
                enhanced=f"[PDF: {fname}]\n{text[:6000]}\n\nStudent question: {question or 'Solve এবং explain করো'}"
                async for evt in SmartTutorAgent().run(enhanced,hist,sid,lang=lang):
                    if evt.get("type")=="token": full+=evt.get("content","")
                    yield f"data: {json.dumps(evt)}\n\n"
            else:
                async for evt in FileAgent().analyze(fpath,fname,q,sid,lang=lang):
                    if evt.get("type")=="token": full+=evt.get("content","")
                    yield f"data: {json.dumps(evt)}\n\n"
            if full:
                mem_add(sid,"user",f"[File: {fname}] {question}")
                mem_add(sid,"assistant",full)
        except Exception as e:
            log.error(f"Upload: {e}",exc_info=True)
            yield f"data: {json.dumps({'type':'error','message':str(e)})}\n\n"
        yield "data: [DONE]\n\n"
    return StreamingResponse(stream(),media_type="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.post("/api/exam/generate")
async def gen_exam(req:ExamRequest):
    prompt=f"""{EXAM_PROMPT}\n\nGenerate complete {req.type_} exam:
Subject: {req.subject} | Class: {req.class_} | Topic: {req.topic or 'Full syllabus'}
Questions: {req.q_count} | Language: {"Bengali (বাংলা)" if req.lang=="bn" else "English"}
Include: header, full questions, answer key."""
    msgs=[{"role":"system","content":EXAM_PROMPT},{"role":"user","content":prompt}]
    full=""
    async def collect():
        nonlocal full
        async for tok in llm(msgs,"general"): full+=tok
    try:
        await collect()
        return JSONResponse({"ok":True,"paper":full,"subject":req.subject,"class_":req.class_})
    except Exception as e:
        return JSONResponse({"ok":False,"error":str(e)},status_code=500)

@app.get("/api/curriculum")
async def curriculum():
    return {"curriculum":{"Primary":{"classes":["Class 1-5"],"subjects":["Bangla","English","Math","Science","Bangladesh Studies"]},
        "JSC":{"classes":["Class 6-8"],"subjects":["Bangla","English","Math","Science","Social Science","ICT"]},
        "SSC":{"classes":["Class 9-10"],"subjects":["Bangla","English","Math","Physics","Chemistry","Biology","Higher Math","Economics","Accounting","ICT"],"groups":["Science","Commerce","Arts"]},
        "HSC":{"classes":["Class 11-12"],"subjects":["Bangla","English","Physics","Chemistry","Biology","Higher Math","Economics","Accounting","ICT"],"groups":["Science","Commerce","Arts"]}}}

@app.get("/api/live-feed")
async def live_feed(category:Optional[str]=None,limit:int=10):
    """Get latest articles from live knowledge engine."""
    return {"items":feed_get(category=category,limit=limit),"category":category}

@app.get("/api/trending")
async def trending():
    """Get what's trending right now."""
    tech=feed_get("tech",5); ai=feed_get("ai",3); news=feed_get("news",5)
    return {"tech":tech,"ai":ai,"world_news":news,"fetched_at":time.time()}

@app.get("/api/sessions")
async def sessions(): return {"sessions":mem_sessions()}

@app.get("/api/sessions/{sid}")
async def get_session(sid:str): return {"session_id":sid,"messages":mem_get(sid,100)}

@app.delete("/api/sessions/{sid}")
async def del_session(sid:str): mem_delete(sid); return {"ok":True}

@app.get("/api/status")
async def status():
    stats=mem_stats()
    with _db() as c:
        last_fetch=c.execute("SELECT ts FROM knowledge_log ORDER BY ts DESC LIMIT 1").fetchone()
    return {"version":"7.0.0","stats":stats,
            "knowledge_engine":{"status":"running","last_update":stats["last_update"]},
            "features":["live_knowledge","multilingual","smart_tutor","vision","hermes_coding","exam_generator"]}

@app.get("/api/tools/weather")
async def w(city:str="Dhaka"): return tool_weather(city) or {"error":"Unavailable"}
@app.get("/api/tools/crypto")
async def c(coins:str="bitcoin,ethereum"): return tool_crypto(coins) or {"error":"Unavailable"}
@app.get("/api/tools/currency")
async def fx(base:str="USD"): return tool_currency(base) or {"error":"Unavailable"}

if __name__=="__main__":
    import uvicorn
    # HuggingFace Spaces uses PORT=7860
    # Local dev uses 8000
    PORT = int(os.getenv("PORT", 7860))
    print()
    print("="*58)
    print("  RUBRA v7 — Always-Updated Intelligence")
    print(f"  Running on port {PORT}")
    print("="*58)
    print()
    uvicorn.run(app, host="0.0.0.0", port=PORT, reload=False)
